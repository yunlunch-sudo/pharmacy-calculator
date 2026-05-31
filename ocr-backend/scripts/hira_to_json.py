"""HIRA 약제급여목록표 Excel → data/drug_db.json 변환.

사용:
    cd ocr-backend
    python scripts/hira_to_json.py <엑셀파일경로> [출력경로]

다운로드:
    https://www.hira.or.kr/bbsDummy.do?pgmid=HIRAA030014050000
    → "약제급여목록 및 급여상한금액표" 최신본 엑셀

매월 1회 갱신 권장. 결과 JSON을 git commit + push 하면 Render가 자동 재배포해
백엔드의 약가가 최신화됨.
"""
import json
import sys
from pathlib import Path

import openpyxl

# CLAUDE.md 기재 컬럼: 제품코드 9번째, 상한금액 14번째 (1-based → 0-based 8/13).
# 헤더 자동탐지로 컬럼 위치 확인 후 폴백.
DEFAULT_CODE_COL = 8
DEFAULT_NAME_COL = 4   # "제품명" 보통 5번째
DEFAULT_PRICE_COL = 13


def find_columns(header_row):
    """헤더 행에서 제품코드/제품명/상한금액 컬럼 위치 찾기."""
    code_idx = name_idx = price_idx = None
    for j, cell in enumerate(header_row):
        s = str(cell or "").strip()
        if not s:
            continue
        if code_idx is None and ("제품코드" in s or "보험코드" in s):
            code_idx = j
        elif name_idx is None and ("제품명" in s or "품명" in s):
            name_idx = j
        elif price_idx is None and "상한금액" in s:
            price_idx = j
    return code_idx, name_idx, price_idx


def main():
    if len(sys.argv) < 2:
        print(__doc__, file=sys.stderr)
        sys.exit(1)

    src = Path(sys.argv[1])
    dst = Path(sys.argv[2]) if len(sys.argv) > 2 else Path(__file__).parent.parent / "data" / "drug_db.json"

    if not src.exists():
        print(f"파일 없음: {src}", file=sys.stderr)
        sys.exit(2)

    print(f"엑셀 로딩: {src}", file=sys.stderr)
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    ws = wb.active

    # 헤더 자동탐지 (제품코드/상한금액이 들어있는 첫 행)
    code_idx = name_idx = price_idx = None
    data_start = 0
    for i, row in enumerate(ws.iter_rows(values_only=True, max_row=20)):
        c, n, p = find_columns(row)
        if c is not None and p is not None:
            code_idx, name_idx, price_idx = c, n, p
            data_start = i + 1
            print(f"헤더 행 {i+1}: 제품코드={c}, 제품명={n}, 상한금액={p}", file=sys.stderr)
            break

    if code_idx is None:
        # 폴백: CLAUDE.md 기준 고정 위치
        code_idx, name_idx, price_idx = DEFAULT_CODE_COL, DEFAULT_NAME_COL, DEFAULT_PRICE_COL
        data_start = 1
        print(f"헤더 자동탐지 실패 → 기본위치 사용 ({code_idx}/{name_idx}/{price_idx})", file=sys.stderr)

    rows = []
    seen = set()
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < data_start:
            continue
        try:
            code_raw = row[code_idx]
            if code_raw is None:
                continue
            code = str(code_raw).strip()
            # 숫자만 (8~9자리). zfill로 9자리 통일.
            if not code.isdigit() or not (7 <= len(code) <= 9):
                continue
            code9 = code.zfill(9)
            if code9 in seen:
                continue
            seen.add(code9)

            name = ""
            if name_idx is not None and name_idx < len(row) and row[name_idx]:
                name = str(row[name_idx]).strip()

            price_raw = row[price_idx]
            if price_raw is None or price_raw == "":
                continue
            try:
                price = int(float(str(price_raw).replace(",", "")))
            except (TypeError, ValueError):
                continue
            if price < 0:
                continue

            rows.append({"code": code9, "name": name, "price": price})
        except (IndexError, TypeError):
            continue

    dst.parent.mkdir(parents=True, exist_ok=True)
    with open(dst, "w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, separators=(",", ":"))
    print(f"✓ {len(rows):,}개 약품 → {dst} ({dst.stat().st_size/1024/1024:.2f} MB)", file=sys.stderr)


if __name__ == "__main__":
    main()
